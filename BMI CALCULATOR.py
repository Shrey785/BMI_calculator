import tkinter as tk
from tkinter import*
from tkinter import messagebox
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

def validateLogin(bmi):
    excel_file = "bmi.xlsx"
    height, age, weight, name = collect_user_data()
    update_excel_sheet(excel_file,height,age,weight,bmi,name)
    return True

def collect_user_data():
    height =  height_entry.get()
    age = Age_entry.get()
    weight = weight_entry.get()
    name = Name_entry.get()
    # result = result_label.get()
    
    return height,age,weight,name

def update_excel_sheet(file_path, height,age,weight,bmi_index, name):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        next_row = sheet.max_row + 1
        if sheet.max_row == 1:
            sheet.cell(row=1, column=1, value="Name")
            sheet.cell(row=1, column=2, value="Age")
            sheet.cell(row=1, column=3, value="Height")
            sheet.cell(row=1, column=4, value="Weight")
            sheet.cell(row=1, column=5, value="Bmi_index")

        sheet.cell(row=next_row, column=1, value=name)
        sheet.cell(row=next_row, column=2, value=age)
        sheet.cell(row=next_row, column=3, value=height)
        sheet.cell(row=next_row, column=4, value=weight)
        sheet.cell(row=next_row, column=5, value=bmi_index)

        workbook.save(file_path)
        print("Data saved successfully!")

    except Exception as e:
        print(f"Error: {e}")

# def update_bmi(target_value, data_to_insert):

#     workbook = openpyxl.load_workbook("bmi.xlsx")
#     worksheet = workbook.active
#     column_letter = 'B'

#     found_row = None
#     print(worksheet[column_letter])
#     for cell in worksheet[column_letter]:
#         if cell.value == target_value:
#             found_row = cell.row
#             break
#     for row_index, (name, age) in enumerate(new_data, start=last_row):
#         sheet[f'A{row_index}'] = name
#         sheet[f'B{row_index}'] = age
#     cell = worksheet.cell(row=found_row, column=openpyxl.utils.column_index_from_string('C'))
#     cell.value = data_to_insert
#     print(cell, cell.value)
#     workbook.save('bmi.xlsx')
#     workbook.close()

ws = Tk()
ws.title('BMI_CALCULATOR')
ws.geometry('750x650')
ws.config(bg='green')


def calc_bmi():
    try:
        height = float((height_entry.get()))/100
        weight = float((weight_entry.get()))
        bmi = float(weight / (height**2))
        bmi = round(bmi, 1)
        bmi_index(bmi)
        validateLogin(bmi)
        plot_graph(weight, height, bmi)

    except ValueError:
        result_label.config("Pls enter valid weight anf height")

def bmi_index(bmi):
    if bmi < 18.5:
        result_label.config(text=f"Your BMI is: {bmi:.2f}, is Underweight")
        # messagebox.show(f'BMI = {bmi} is Underweight')
    elif (bmi > 18.5) and (bmi < 24.9):
        result_label.config(text=f"Your BMI is: {bmi:.2f}, is Normal")
        # messagebox.showinfo(f'BMI = {bmi} is Normal')
    elif (bmi > 24.9) and (bmi < 29.9):
        result_label.config(text=f"Your BMI is: {bmi:.2f}, is overweight")
        # messagebox.showinfo(f'BMI = {bmi} is Overweight')
    else:
        result_label.config(text=f"Your BMI is: {bmi:.2f}, is Obesity")
        # messagebox.showinfo('Obesity')  
var = IntVar()

Name_label = Label(text='Name', font=('Arial', 10, 'bold')).pack(pady=0)
Name_entry = tk.Entry(ws)
Name_entry.pack(pady=2)

Age_label = Label(text='Age', font=('Arial', 10, 'bold')).pack(pady=0)
Age_entry = tk.Entry(ws)
Age_entry.pack(pady=2)

tk.Label(ws, text="Weight (kg):", font=('Arial', 10, 'bold')).pack(pady=0)
weight_entry = tk.Entry(ws)
weight_entry.pack(pady=2)

tk.Label(ws, text="Height (cm):",font=('Arial', 10, 'bold')).pack(pady=0)
height_entry = tk.Entry(ws)
height_entry.pack(pady=2)

calculate_button = tk.Button(ws, text="Calculate BMI", command=calc_bmi)
calculate_button.pack(pady=10)

result_label = tk.Label(ws, text="")
result_label.pack(pady=10)

# update_bmi(Age_entry.get(), weight_entry.get(), height_entry.get(), bmi_index.get())

def plot_graph(weight, height, bmi):
    weight = np.array([20, 40, 60, 80, 100, 120, 140, 160, 180, 200])

    height = np.array([150, 155, 160, 165, 170, 175, 180, 185, 190.195, 200])
    print(weight, height, bmi)
    plt.plot(weight, height, bmi)

    plt.xlabel('weight')
    plt.ylabel('height')

    plt.title("BMI GRAPH")

    plt.show()
ws.mainloop()